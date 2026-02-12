import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart
from openpyxl.chart.reference import Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# 讀取 CSV 檔案
csv_path = '/workspaces/StarPilot/examples/02_數據分析/測試資料/sales_data.csv'
df = pd.read_csv(csv_path)

# 數據分析
print("=" * 60)
print("銷售資料分析")
print("=" * 60)

# 基本統計
print(f"\n1. 基本統計信息:")
print(f"   - 總記錄數: {len(df)}")
print(f"   - 日期範圍: {df['日期'].min()} 到 {df['日期'].max()}")
print(f"   - 總銷售金額: NT${df['金額'].sum():,.0f}")
print(f"   - 總銷售數量: {df['數量'].sum()} 個")

# 按產品分析
print(f"\n2. 產品銷售統計:")
product_sales = df.groupby('產品').agg(
    {'金額': 'sum', '數量': 'sum'}).sort_values('金額', ascending=False)
print(product_sales)

# 按地區分析
print(f"\n3. 地區銷售統計:")
region_sales = df.groupby('地區').agg(
    {'金額': 'sum', '數量': 'sum'}).sort_values('金額', ascending=False)
print(region_sales)

# 按業務分析
print(f"\n4. 業務員銷售統計:")
staff_sales = df.groupby('業務').agg(
    {'金額': 'sum', '數量': 'sum'}).sort_values('金額', ascending=False)
print(staff_sales)

# 按月份分析
df['月份'] = pd.to_datetime(df['日期']).dt.to_period('M')
print(f"\n5. 月份銷售統計:")
monthly_sales = df.groupby('月份').agg({'金額': 'sum', '數量': 'sum'})
print(monthly_sales)

# 創建 Excel 檔案
output_path = '/workspaces/StarPilot/examples/02_數據分析/銷售分析報告.xlsx'
wb = Workbook()
wb.remove(wb.active)

# 設定樣式
header_fill = PatternFill(start_color="4472C4",
                          end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 1. 概況工作表
ws_overview = wb.create_sheet("概況", 0)
ws_overview['A1'] = "銷售資料分析報告"
ws_overview['A1'].font = Font(size=16, bold=True)
ws_overview.merge_cells('A1:B1')

overview_data = [
    ("總銷售金額", f"NT${df['金額'].sum():,.0f}"),
    ("總銷售數量", f"{df['數量'].sum()} 個"),
    ("銷售記錄數", f"{len(df)} 筆"),
    ("日期範圍", f"{df['日期'].min()} ~ {df['日期'].max()}"),
    ("產品種類", f"{df['產品'].nunique()} 種"),
    ("銷售地區", f"{df['地區'].nunique()} 個"),
    ("業務人員", f"{df['業務'].nunique()} 人"),
]

for idx, (label, value) in enumerate(overview_data, 3):
    ws_overview[f'A{idx}'] = label
    ws_overview[f'B{idx}'] = value
    ws_overview[f'A{idx}'].font = Font(bold=True)
    ws_overview.column_dimensions['A'].width = 15
    ws_overview.column_dimensions['B'].width = 20

# 2. 產品銷售分析工作表
ws_product = wb.create_sheet("產品銷售", 1)
ws_product['A1'] = "產品銷售統計"
ws_product['A1'].font = Font(size=14, bold=True)
ws_product.merge_cells('A1:C1')

# 寫入標題
headers = ['產品', '銷售金額', '銷售數量']
for col, header in enumerate(headers, 1):
    cell = ws_product.cell(row=3, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border

# 寫入產品數據
product_data = df.groupby('產品').agg(
    {'金額': 'sum', '數量': 'sum'}).sort_values('金額', ascending=False)
for idx, (product, row) in enumerate(product_data.iterrows(), 4):
    ws_product[f'A{idx}'] = product
    ws_product[f'B{idx}'] = row['金額']
    ws_product[f'C{idx}'] = row['數量']
    ws_product[f'B{idx}'].number_format = '#,##0'

ws_product.column_dimensions['A'].width = 12
ws_product.column_dimensions['B'].width = 15
ws_product.column_dimensions['C'].width = 12

# 產品銷售柱狀圖
chart1 = BarChart()
chart1.type = "col"
chart1.title = "產品銷售金額"
chart1.y_axis.title = "金額 (NT$)"
chart1.x_axis.title = "產品"
data = Reference(ws_product, min_col=2, min_row=3, max_row=3+len(product_data))
cats = Reference(ws_product, min_col=1, min_row=4, max_row=3+len(product_data))
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
ws_product.add_chart(chart1, "A12")

# 3. 地區銷售分析工作表
ws_region = wb.create_sheet("地區銷售", 2)
ws_region['A1'] = "地區銷售統計"
ws_region['A1'].font = Font(size=14, bold=True)
ws_region.merge_cells('A1:C1')

# 寫入標題
for col, header in enumerate(headers, 1):
    cell = ws_region.cell(row=3, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border

# 寫入地區數據
region_data = df.groupby('地區').agg(
    {'金額': 'sum', '數量': 'sum'}).sort_values('金額', ascending=False)
for idx, (region, row) in enumerate(region_data.iterrows(), 4):
    ws_region[f'A{idx}'] = region
    ws_region[f'B{idx}'] = row['金額']
    ws_region[f'C{idx}'] = row['數量']
    ws_region[f'B{idx}'].number_format = '#,##0'

ws_region.column_dimensions['A'].width = 12
ws_region.column_dimensions['B'].width = 15
ws_region.column_dimensions['C'].width = 12

# 地區銷售圓餅圖
chart2 = PieChart()
chart2.title = "各地區銷售佔比"
data = Reference(ws_region, min_col=2, min_row=3, max_row=3+len(region_data))
cats = Reference(ws_region, min_col=1, min_row=4, max_row=3+len(region_data))
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(cats)
ws_region.add_chart(chart2, "A12")

# 4. 業務銷售分析工作表
ws_staff = wb.create_sheet("業務銷售", 3)
ws_staff['A1'] = "業務員銷售統計"
ws_staff['A1'].font = Font(size=14, bold=True)
ws_staff.merge_cells('A1:C1')

# 寫入標題
for col, header in enumerate(headers, 1):
    cell = ws_staff.cell(row=3, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border

# 寫入業務數據
staff_data = df.groupby('業務').agg(
    {'金額': 'sum', '數量': 'sum'}).sort_values('金額', ascending=False)
for idx, (staff, row) in enumerate(staff_data.iterrows(), 4):
    ws_staff[f'A{idx}'] = staff
    ws_staff[f'B{idx}'] = row['金額']
    ws_staff[f'C{idx}'] = row['數量']
    ws_staff[f'B{idx}'].number_format = '#,##0'

ws_staff.column_dimensions['A'].width = 12
ws_staff.column_dimensions['B'].width = 15
ws_staff.column_dimensions['C'].width = 12

# 業務銷售柱狀圖
chart3 = BarChart()
chart3.type = "col"
chart3.title = "業務員銷售金額"
chart3.y_axis.title = "金額 (NT$)"
chart3.x_axis.title = "業務員"
data = Reference(ws_staff, min_col=2, min_row=3, max_row=3+len(staff_data))
cats = Reference(ws_staff, min_col=1, min_row=4, max_row=3+len(staff_data))
chart3.add_data(data, titles_from_data=True)
chart3.set_categories(cats)
ws_staff.add_chart(chart3, "A12")

# 5. 月份趨勢分析工作表
ws_monthly = wb.create_sheet("月份趨勢", 4)
ws_monthly['A1'] = "月份銷售趨勢"
ws_monthly['A1'].font = Font(size=14, bold=True)
ws_monthly.merge_cells('A1:C1')

# 寫入標題
monthly_headers = ['月份', '銷售金額', '銷售數量']
for col, header in enumerate(monthly_headers, 1):
    cell = ws_monthly.cell(row=3, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border

# 寫入月份數據
monthly_data = df.groupby('月份').agg({'金額': 'sum', '數量': 'sum'})
for idx, (month, row) in enumerate(monthly_data.iterrows(), 4):
    ws_monthly[f'A{idx}'] = str(month)
    ws_monthly[f'B{idx}'] = row['金額']
    ws_monthly[f'C{idx}'] = row['數量']
    ws_monthly[f'B{idx}'].number_format = '#,##0'

ws_monthly.column_dimensions['A'].width = 12
ws_monthly.column_dimensions['B'].width = 15
ws_monthly.column_dimensions['C'].width = 12

# 月份銷售趨勢折線圖
chart4 = LineChart()
chart4.title = "月份銷售趨勢"
chart4.y_axis.title = "金額 (NT$)"
chart4.x_axis.title = "月份"
data = Reference(ws_monthly, min_col=2, min_row=3, max_row=3+len(monthly_data))
cats = Reference(ws_monthly, min_col=1, min_row=4, max_row=3+len(monthly_data))
chart4.add_data(data, titles_from_data=True)
chart4.set_categories(cats)
ws_monthly.add_chart(chart4, "A12")

# 保存 Excel 檔案
wb.save(output_path)
print("\n" + "=" * 60)
print(f"✓ Excel 報告已生成: {output_path}")
print("=" * 60)
